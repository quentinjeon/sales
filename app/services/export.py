# -*- coding: utf-8 -*-
"""결과 엑셀 내보내기 — PRD §8

시트 6장. 화면으로 못 보는 전체 목록과, 다른 사람에게 넘길 때 쓴다.

    01_요약  02_제품별  03_채널별  04_채널x제품  05_적자딜  06_미매핑
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import Channel, SalesLine
from app.services import report as rp

HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
MONEY = "#,##0"
PCT = "0.00%"


def _sheet(wb: Workbook, title: str, headers: list[str], widths: dict | None = None):
    ws = wb.active if wb.sheetnames == ["Sheet"] else wb.create_sheet()
    ws.title = title
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = (widths or {}).get(i, 14)
    ws.freeze_panes = "A2"
    return ws


def _fmt(ws, cols: str, number_format: str) -> None:
    for col in cols:
        for cell in ws[col][1:]:
            cell.number_format = number_format


def build(db: Session, period: str, ptype: str = "MONTH") -> Workbook:
    """기간 하나에 대한 결과 워크북을 만든다."""
    wb = Workbook()
    t = rp.totals(db, period, ptype)
    q = rp.quality(db, period, ptype)

    # 01 요약
    ws = _sheet(wb, "01_요약", ["항목", "값"], {1: 24, 2: 20})
    for k, v in [("기간", period), ("주문 건수", t.orders), ("판매 수량", t.qty),
                 ("총매출", t.revenue), ("채널 수수료", -t.fee), ("원가", -t.cogs),
                 ("물류비", -t.logistics), ("기여이익", t.profit)]:
        ws.append([k, v])
    ws.append(["마진율", t.margin])
    ws.append(["미매핑 건수", q.unmapped_count])
    ws.append(["미매핑 금액", q.unmapped_amount])
    for r in range(4, 10):
        ws.cell(row=r, column=2).number_format = MONEY
    ws.cell(row=10, column=2).number_format = PCT
    ws.cell(row=12, column=2).number_format = MONEY

    # 02 제품별
    ws = _sheet(wb, "02_제품별",
                ["제품코드", "제품명", "수량", "매출", "수수료", "원가", "물류비",
                 "기여이익", "마진율"], {1: 12, 2: 34})
    for r in rp.by_product(db, period, ptype):
        ws.append([r.key, r.name, r.qty, r.revenue, r.fee, r.cogs, r.logistics,
                   r.profit, r.margin])
    _fmt(ws, "CDEFGH", MONEY)
    _fmt(ws, "I", PCT)

    # 03 채널별
    ws = _sheet(wb, "03_채널별",
                ["채널코드", "채널명", "주문건수", "매출", "수수료율", "수수료", "원가",
                 "물류비", "기여이익", "마진율"], {1: 12, 2: 20})
    for r in rp.by_channel(db, period, ptype):
        ws.append([r.key, r.name, r.orders, r.revenue, r.extra.get("fee_rate"),
                   r.fee, r.cogs, r.logistics, r.profit, r.margin])
    _fmt(ws, "CDFGHI", MONEY)
    _fmt(ws, "EJ", PCT)

    # 04 채널 × 제품 / 05 적자 딜
    deal_head = ["채널", "제품(딜)", "티어", "판매가", "주문건수", "매출", "기여이익", "마진율"]
    for title, rows in (("04_채널x제품", rp.by_deal(db, period, ptype, order="profit")),
                        ("05_적자딜", rp.loss_deals(db, period, ptype))):
        ws = _sheet(wb, title, deal_head, {1: 20, 2: 34})
        for r in rows:
            ws.append([r.sub, r.name, r.extra["tier"], r.extra["price"], r.orders,
                       r.revenue, r.profit, r.margin])
        _fmt(ws, "DEFG", MONEY)
        _fmt(ws, "H", PCT)

    # 06 미매핑 — 조용히 버리지 않는다는 약속을 파일로도 지킨다
    ws = _sheet(wb, "06_미매핑", ["채널", "상품명", "옵션명", "건수", "금액"],
                {1: 20, 2: 46, 3: 40})
    names = {c.id: c.name for c in db.scalars(select(Channel)).all()}
    col = SalesLine.period_month if ptype == "MONTH" else SalesLine.period_week
    for ch, prod, opt, cnt, amt in db.execute(
            select(SalesLine.channel_id, SalesLine.raw_product_name,
                   SalesLine.raw_option_name, func.count(),
                   func.sum(SalesLine.gross_amount))
            .where(SalesLine.map_status == "UNMAPPED", col == period)
            .group_by(SalesLine.channel_id, SalesLine.raw_product_name,
                      SalesLine.raw_option_name)
            .order_by(func.sum(SalesLine.gross_amount).desc())).all():
        ws.append([names.get(ch, ch), prod, opt, cnt, amt])
    _fmt(ws, "E", MONEY)

    return wb


def to_bytes(db: Session, period: str, ptype: str = "MONTH") -> bytes:
    buf = BytesIO()
    build(db, period, ptype).save(buf)
    return buf.getvalue()
