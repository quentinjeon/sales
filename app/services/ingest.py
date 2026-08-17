# -*- coding: utf-8 -*-
"""엑셀 적재 — PRD §4

    01_상품정보.xlsx   → sku · sku_cost · deal · deal_component
    02_채널정보.xlsx   → channel · channel_fee · channel_logistics
    03_매출_YYYY-MM.xlsx → upload_batch · upload_file · sales_line (+ 매핑 · 계산)

채널마다 컬럼 이름이 다르다(PRD §4.3). 그래서 컬럼을 이름으로 직접 찾지 않고
'논리 필드 → 후보 키워드' 표로 찾는다. 채널이 어드민을 개편해 컬럼명을 바꿔도
키워드만 추가하면 된다.

찾지 못한 필수 컬럼은 예외로 알린다 — 조용히 빈 값으로 채우면 매출이 증발한다.
"""
from __future__ import annotations

import hashlib
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import (
    Channel, ChannelFee, ChannelLogistics, Deal, DealComponent, SalesLine, Sku, SkuCost,
    UploadBatch, UploadFile,
)
from app.seed.data import TIERS
from app.services import mapping as mp
from app.services.calc import CalcError, apply_to_line, clear_calc
from app.services.report import period_keys

NOTE_SHEETS = {"읽어주세요", "안내", "readme"}

# 논리 필드 → 컬럼명에 들어갈 법한 키워드 (긴 것부터 검사한다)
SALES_COLUMNS = {
    "주문번호": ["주문번호", "주문id", "주문 id", "orderid", "order_no"],
    "주문일":   ["주문일자", "구매확정일", "결제일", "정산일", "매출일", "주문일", "date"],
    "상품명":   ["노출상품명", "주문상품", "상품명", "품목명", "품목", "product"],
    "옵션명":   ["등록옵션명", "선택옵션", "단품명", "옵션명", "옵션", "option"],
    "수량":     ["판매수량", "주문수량", "수량", "개수", "qty"],
    "판매금액": ["판매금액", "결제금액", "상품금액", "정산금액", "매출액", "판매가"],
    "수수료":   ["서비스이용료", "판매수수료", "매입수수료", "수수료", "이용료"],
    "취소여부": ["취소여부", "주문상태", "취소", "반품", "상태"],
}
SALES_REQUIRED = ["주문일", "상품명", "수량", "판매금액"]

CANCEL_WORDS = ("취소", "반품", "환불", "cancel", "return")


class IngestError(Exception):
    """파일을 쓸 수 없다 — 필수 컬럼 누락, 채널 미등록 등."""


@dataclass
class SalesResult:
    batch_id: int = 0
    period: str = ""
    files: int = 0
    lines: int = 0
    mapped: int = 0
    auto_mapped: int = 0
    unmapped: int = 0
    calculated: int = 0
    calc_error: int = 0
    skipped_sheets: list[str] = field(default_factory=list)
    gross_sum: int = 0
    dup_in_file: int = 0        # 같은 파일 안에서 중복된 행
    dup_existing: int = 0       # 이미 적재된 주문과 겹친 행
    replaced_lines: int = 0     # 교체 모드에서 지운 기존 행


# ── 시트 읽기 ───────────────────────────────────────────────────────────

def _rows(ws) -> list[list]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _header_index(headers: list, spec: dict[str, list[str]]) -> dict[str, int]:
    """헤더 행에서 논리 필드별 컬럼 위치를 찾는다."""
    norm = [str(h).strip().lower().replace(" ", "") if h is not None else "" for h in headers]
    found: dict[str, int] = {}
    for field_name, keywords in spec.items():
        for kw in keywords:
            k = kw.replace(" ", "")
            for i, h in enumerate(norm):
                if h and k in h and i not in found.values():
                    found[field_name] = i
                    break
            if field_name in found:
                break
    return found


def _int(v, default: int | None = 0) -> int | None:
    if v is None or v == "":
        return default
    if isinstance(v, str):
        v = v.replace(",", "").replace("원", "").strip()
        if not v:
            return default
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return default


def _date(v) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if v is None:
        return None
    s = str(v).strip()[:10].replace("/", "-").replace(".", "-")
    try:
        return date.fromisoformat(s).isoformat()
    except ValueError:
        return None


# ── ① 상품정보 ─────────────────────────────────────────────────────────

def load_products(db: Session, path: str | Path, effective_from: str = "2026-01-01") -> dict:
    """상품정보 엑셀 → sku · sku_cost · deal(3티어) · deal_component.

    이미 있는 제품은 건너뛴다. 매입가를 바꿀 때는 이력 행을 따로 추가해야 하며
    (`sku_cost`), 여기서 덮어쓰지 않는다 — 과거 손익이 소급 변형되면 안 된다.
    """
    ws = load_workbook(path, data_only=True)["상품정보"]
    rows = _rows(ws)
    if not rows:
        raise IngestError(f"빈 파일입니다: {path}")
    idx = _header_index(rows[0], {
        "제품코드": ["제품코드", "sku"], "제품명": ["제품명", "상품명"],
        "규격": ["규격", "spec"], "매입가": ["매입가", "원가", "매입"],
        "구성수량": ["구성수량", "구성", "pack"],
        "정상가": ["정상가"], "일반행사가": ["일반행사가", "행사가"], "특가": ["특가"],
    })
    need = ["제품코드", "제품명", "매입가", "구성수량", "정상가", "일반행사가", "특가"]
    missing = [k for k in need if k not in idx]
    if missing:
        raise IngestError(f"상품정보 필수 컬럼 누락: {', '.join(missing)}")

    n_sku = n_deal = 0
    price_cols = {"NORMAL": idx["정상가"], "EVENT": idx["일반행사가"], "SPECIAL": idx["특가"]}
    for r in rows[1:]:
        sku_id = str(r[idx["제품코드"]] or "").strip()
        if not sku_id:
            continue
        cost, pack = _int(r[idx["매입가"]]), _int(r[idx["구성수량"]])
        if not cost or not pack:
            raise IngestError(f"{sku_id}: 매입가·구성수량이 비었습니다")
        if not db.get(Sku, sku_id):
            db.add(Sku(id=sku_id, name=str(r[idx["제품명"]]).strip(),
                       spec=str(r[idx["규격"]]).strip() if "규격" in idx else None,
                       brand=None))
            db.add(SkuCost(sku_id=sku_id, unit_cost=cost, effective_from=effective_from,
                           source=Path(path).name))
            n_sku += 1
        for tier, _ in TIERS:
            price = _int(r[price_cols[tier]])
            if not price:
                continue
            deal_id = f"{sku_id}-X{pack}-{tier}"
            if db.get(Deal, deal_id):
                continue
            db.add(Deal(id=deal_id, channel_id=None, primary_sku_id=sku_id,
                        label=f"{str(r[idx['제품명']]).strip()} ×{pack}",
                        tier=tier, price=price, effective_from=effective_from))
            db.add(DealComponent(deal_id=deal_id, sku_id=sku_id, qty=pack))
            n_deal += 1
    db.flush()
    return {"skus": n_sku, "deals": n_deal}


# ── ② 채널정보 ─────────────────────────────────────────────────────────

def load_channels(db: Session, path: str | Path, effective_from: str = "2026-01-01") -> dict:
    """채널정보 엑셀 → channel · channel_fee · channel_logistics."""
    ws = load_workbook(path, data_only=True)["채널정보"]
    rows = _rows(ws)
    idx = _header_index(rows[0], {
        "채널코드": ["채널코드"], "채널명": ["채널명"], "그룹": ["그룹"],
        "운영상태": ["운영상태", "상태"], "수수료율": ["수수료율"],
        "요율근거": ["요율근거", "근거"], "배송주체": ["배송주체"],
        "물류비": ["물류비"], "물류비추정": ["물류비추정", "추정"],
        "정산기준일": ["정산기준일"], "비고": ["비고"],
    })
    missing = [k for k in ("채널코드", "채널명", "수수료율", "배송주체") if k not in idx]
    if missing:
        raise IngestError(f"채널정보 필수 컬럼 누락: {', '.join(missing)}")

    n = 0
    for r in rows[1:]:
        cid = str(r[idx["채널코드"]] or "").strip()
        if not cid or db.get(Channel, cid):
            continue
        ship = str(r[idx["배송주체"]] or "SELF").strip()
        logi_raw = r[idx["물류비"]] if "물류비" in idx else None
        bears = isinstance(logi_raw, str) and "채널" in logi_raw
        db.add(Channel(
            id=cid, name=str(r[idx["채널명"]]).strip(),
            group_name=str(r[idx["그룹"]] or "기타").strip() if "그룹" in idx else "기타",
            fee_base="PURCHASE" if ship == "CONSIGNMENT" else "PRICE",
            ship_owner=ship,
            settle_date_src=str(r[idx["정산기준일"]] or "CONFIRM").strip()
            if "정산기준일" in idx else "CONFIRM",
            vat_basis="EXCLUDED",
            status=str(r[idx["운영상태"]] or "WAITING").strip() if "운영상태" in idx else "WAITING",
            sort_order=n + 1,
            note=str(r[idx["비고"]]).strip() if "비고" in idx and r[idx["비고"]] else None,
        ))
        rate = r[idx["수수료율"]]
        if rate not in (None, ""):
            db.add(ChannelFee(
                channel_id=cid, fee_rate=float(rate), effective_from=effective_from,
                source=str(r[idx["요율근거"]] or "ESTIMATE").strip()
                if "요율근거" in idx and r[idx["요율근거"]] else "ESTIMATE"))
        db.add(ChannelLogistics(
            channel_id=cid,
            model="CHANNEL_BEARS" if bears else "FLAT",
            flat_amount=None if bears else _int(logi_raw, 0),
            is_estimate=_int(r[idx["물류비추정"]], 0) if "물류비추정" in idx else 0,
            effective_from=effective_from))
        n += 1
    db.flush()
    return {"channels": n}


# ── ③ 매출 원장 ────────────────────────────────────────────────────────

def file_digest(path: str | Path) -> str:
    return hashlib.sha256(Path(path).read_bytes()).hexdigest()


def find_duplicate_file(db: Session, digest: str) -> UploadFile | None:
    """같은 내용의 파일이 이미 적재되어 있는가 (§7.6 중복 방지)."""
    return db.scalar(
        select(UploadFile).where(UploadFile.sha256 == digest)
        .order_by(UploadFile.id.desc()))


def _existing_keys(db: Session, channel_id: str) -> set[tuple]:
    """이 채널에 이미 적재된 주문 키. 같은 주문번호는 같은 주문이다."""
    return set(db.execute(
        select(SalesLine.order_no, SalesLine.order_line_no, SalesLine.raw_option_name)
        .where(SalesLine.channel_id == channel_id)).all())


def wipe_period(db: Session, channel_id: str, period: str) -> int:
    """한 채널의 한 기간 데이터를 지운다 (교체 업로드용). 지운 행 수를 돌려준다."""
    fids = set(db.scalars(
        select(SalesLine.file_id).where(SalesLine.channel_id == channel_id,
                                        SalesLine.period_month == period)).all())
    n = db.query(SalesLine).filter(
        SalesLine.channel_id == channel_id,
        SalesLine.period_month == period).delete(synchronize_session=False)
    for fid in fids:                     # 라인이 하나도 안 남은 파일 행은 같이 정리
        if not db.scalar(select(func.count()).select_from(SalesLine)
                         .where(SalesLine.file_id == fid)):
            db.query(UploadFile).filter(UploadFile.id == fid).delete(
                synchronize_session=False)
    db.flush()
    return n


def load_sales(db: Session, path: str | Path, *, period: str | None = None,
               uploaded_by: str | None = None, replace: bool = False) -> SalesResult:
    """매출 엑셀(채널별 시트) → sales_line, 그리고 매핑 → 손익 계산까지.

    중복 방지 3단계 (PRD §7.6):
      ① 같은 파일(sha256)이 이미 적재되어 있으면 거부한다
      ② 이미 적재된 주문번호와 겹치는 행은 건너뛴다
      ③ replace=True 면 해당 채널×기간을 지우고 새로 넣는다

    ①②가 없으면 같은 파일을 두 번 올릴 때 매출이 그대로 두 배가 된다.
    """
    path = Path(path)
    digest = file_digest(path)

    if not replace:
        dup = find_duplicate_file(db, digest)
        if dup is not None:
            total = db.scalar(select(func.coalesce(func.sum(UploadFile.row_count), 0))
                              .where(UploadFile.sha256 == digest)) or 0
            raise IngestError(
                f"이미 올린 파일입니다 — '{dup.filename}' "
                f"({(dup.created_at or '')[:10]} 적재 · {total:,}건). "
                f"같은 기간을 다시 계산하려면 '기존 데이터 교체'를 체크하고 올리세요.")

    wb = load_workbook(path, data_only=True)
    by_name = {c.name: c.id for c in db.scalars(select(Channel)).all()}
    by_id = {c.id: c.id for c in db.scalars(select(Channel)).all()}

    res = SalesResult(period=period or "")
    batch = UploadBatch(period_type="MONTH", period_key=period or "", status="UPLOADING",
                        uploaded_by=uploaded_by)
    db.add(batch)
    db.flush()
    res.batch_id = batch.id

    # ── 1단계: 파싱만 한다 (DB 쓰기 전에 기간·중복을 먼저 확정해야 한다) ──
    parsed: list[tuple[str, list[dict]]] = []
    for ws in wb.worksheets:
        title = ws.title.strip()
        if title.lower() in NOTE_SHEETS:
            continue
        channel_id = by_name.get(title) or by_id.get(title.upper())
        if channel_id is None:
            res.skipped_sheets.append(f"{title} (채널 미등록)")
            continue

        rows = _rows(ws)
        if len(rows) < 2:
            res.skipped_sheets.append(f"{title} (데이터 없음)")
            continue
        idx = _header_index(rows[0], SALES_COLUMNS)
        missing = [k for k in SALES_REQUIRED if k not in idx]
        if missing:
            raise IngestError(f"[{title}] 필수 컬럼 누락: {', '.join(missing)}")

        seen: set[tuple] = set()
        recs: list[dict] = []
        for i, r in enumerate(rows[1:], start=2):
            on = _date(r[idx["주문일"]])
            product = str(r[idx["상품명"]] or "").strip()
            qty = _int(r[idx["수량"]], 0)
            if not on or not product or not qty:
                continue
            option = str(r[idx["옵션명"]] or "").strip() if "옵션명" in idx else ""
            order_no = (str(r[idx["주문번호"]]).strip() if "주문번호" in idx
                        and r[idx["주문번호"]] else f"{channel_id}-{i:07d}")
            key = (order_no, "1", option)
            if key in seen:                          # ① 같은 파일 안의 중복 행
                res.dup_in_file += 1
                continue
            seen.add(key)

            cancelled = 0
            if "취소여부" in idx and r[idx["취소여부"]]:
                cancelled = int(any(w in str(r[idx["취소여부"]]).lower() for w in CANCEL_WORDS))
            pm, pw = period_keys(date.fromisoformat(on))
            recs.append({
                "key": key, "order_no": order_no, "order_date": on,
                "product": product, "option": option, "qty": qty,
                "amount": _int(r[idx["판매금액"]], 0),
                "fee_actual": _int(r[idx["수수료"]], None) if "수수료" in idx else None,
                "cancelled": cancelled, "pm": pm, "pw": pw,
            })
        if recs:
            parsed.append((channel_id, recs))

    if not parsed:
        raise IngestError("읽을 수 있는 채널 시트가 없습니다. 시트명이 채널명과 같아야 합니다.")

    # ── 2단계: 교체 모드면 해당 채널×기간을 먼저 비운다 ──
    if replace:
        for channel_id, recs in parsed:
            for pm in {r["pm"] for r in recs}:
                res.replaced_lines += wipe_period(db, channel_id, pm)

    # ── 3단계: 이미 적재된 주문과 겹치는 행을 걸러내고 적재한다 ──
    all_lines: list[SalesLine] = []
    for channel_id, recs in parsed:
        existing = _existing_keys(db, channel_id)
        fresh = [r for r in recs if r["key"] not in existing]
        res.dup_existing += len(recs) - len(fresh)
        if not fresh:
            res.skipped_sheets.append(f"{channel_id} (전부 이미 적재됨)")
            continue

        f = UploadFile(batch_id=batch.id, channel_id=channel_id, filename=path.name,
                       stored_path=str(path), sha256=digest, status="PARSED")
        db.add(f)
        db.flush()

        gross = 0
        for r in fresh:
            db.add(ln := SalesLine(
                file_id=f.id, channel_id=channel_id, order_no=r["order_no"],
                order_line_no="1", order_date=r["order_date"],
                raw_product_name=r["product"], raw_option_name=r["option"],
                qty=r["qty"], gross_amount=r["amount"], fee_actual=r["fee_actual"],
                is_cancelled=r["cancelled"], period_month=r["pm"], period_week=r["pw"],
                map_status="UNMAPPED"))
            all_lines.append(ln)
            gross += r["amount"]

        f.row_count, f.gross_sum = len(fresh), gross
        res.files += 1
        res.lines += len(fresh)
        res.gross_sum += gross
        db.flush()

    if not all_lines:
        raise IngestError(
            f"새로 적재할 주문이 없습니다 — {res.dup_existing:,}건이 이미 들어와 있습니다.")

    stat = mp.map_lines(db, all_lines)
    res.mapped, res.auto_mapped, res.unmapped = stat["mapped"], stat["auto"], stat["unmapped"]

    for ln in all_lines:
        if ln.map_status != "MAPPED":
            clear_calc(ln)
            continue
        try:
            apply_to_line(db, ln, db.get(Deal, ln.deal_id))
            res.calculated += 1
        except CalcError as e:
            ln.map_status, ln.map_note = "UNMAPPED", str(e)
            clear_calc(ln)
            res.calc_error += 1
            res.unmapped += 1
            res.mapped -= 1

    if not res.period and all_lines:
        res.period = max(ln.period_month for ln in all_lines)
    batch.period_key = res.period
    batch.status = "CALCULATED"
    batch.calculated_at = datetime.now().isoformat(timespec="seconds")
    db.flush()
    return res
