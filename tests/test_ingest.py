# -*- coding: utf-8 -*-
"""엑셀 적재 테스트 — PRD §4 · §9 검수 기준

T-I1 이 가장 중요하다. 실습 엑셀 3장을 빈 DB에 넣었을 때
PRD §9 의 검수 기준 숫자가 그대로 나와야 한다.
어댑터·매핑·계산 중 하나라도 틀어지면 여기가 깨진다.
"""
from __future__ import annotations

from pathlib import Path

import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.models import Base, Channel, Deal, Sku
from app.services import ingest as ing
from app.services import report as rp

DATA = Path(__file__).resolve().parent.parent / "docs" / "lecture" / "실습데이터"
PERIOD = "2026-07"

# PRD §9 검수 기준
EXPECTED_TOTALS = {
    "orders": 3707, "revenue": 77_667_300, "fee": 8_711_768,
    "cogs": 52_072_450, "logistics": 10_928_600, "profit": 5_954_482,
}
EXPECTED_UNMAPPED = (7, 214_300)
EXPECTED_CHANNELS = {                       # 채널코드 → (주문, 매출, 기여이익)
    "HANBIT":   (1255, 26_728_500, 2_483_980),
    "SWIFT_FF": (962, 22_497_800, 1_780_042),
    "MALL21":   (769, 14_613_100, 868_547),
    "GOODMKT":  (570, 10_562_000, 652_600),
    "BIDNOW":   (151, 3_265_900, 169_313),
}

pytestmark = pytest.mark.skipif(
    not (DATA / "01_상품정보.xlsx").exists(),
    reason="실습 엑셀이 없습니다 — python docs/lecture/실습데이터_생성.py 로 생성하세요")


@pytest.fixture(scope="module")
def db():
    """빈 DB에 엑셀 3장만 넣는다. 시드는 쓰지 않는다."""
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as s:
        ing.load_products(s, DATA / "01_상품정보.xlsx")
        ing.load_channels(s, DATA / "02_채널정보.xlsx")
        s.result = ing.load_sales(s, DATA / f"03_매출_{PERIOD}.xlsx", period=PERIOD)
        yield s


# ── T-I1 · 검수 기준 재현 ────────────────────────────────────────────────

def test_i1_totals_match_prd(db):
    """T-I1 — 엑셀만으로 PRD §9 의 전체 합계가 재현된다."""
    t = rp.totals(db, PERIOD)
    got = {"orders": t.orders, "revenue": t.revenue, "fee": t.fee,
           "cogs": t.cogs, "logistics": t.logistics, "profit": t.profit}
    assert got == EXPECTED_TOTALS


def test_i1b_margin_rate(db):
    assert rp.totals(db, PERIOD).margin == pytest.approx(0.0767, abs=1e-4)


@pytest.mark.parametrize("cid,expected", EXPECTED_CHANNELS.items())
def test_i1c_by_channel(db, cid, expected):
    row = next(r for r in rp.by_channel(db, PERIOD) if r.key == cid)
    assert (row.orders, row.revenue, row.profit) == expected


# ── T-I2 · 마스터 적재 ───────────────────────────────────────────────────

def test_i2_masters_loaded(db):
    from sqlalchemy import func, select
    assert db.scalar(select(func.count()).select_from(Sku)) == 23
    assert db.scalar(select(func.count()).select_from(Deal)) == 69     # 23 × 3티어
    assert db.scalar(select(func.count()).select_from(Channel)) == 14


def test_i2b_pack_qty_survives_the_round_trip(db):
    """구성수량이 딜 구성으로 살아있어야 한다 — 여기가 죽으면 원가가 수백 배 작아진다."""
    deal = db.get(Deal, "CUP1000-X1000-NORMAL")
    assert deal is not None
    assert sum(c.qty for c in deal.components) == 1000


# ── T-I3 · 미매핑은 버리지 않는다 ────────────────────────────────────────

def test_i3_unmapped_is_kept(db):
    q = rp.quality(db, PERIOD)
    assert (q.unmapped_count, q.unmapped_amount) == EXPECTED_UNMAPPED


def test_i3b_file_total_equals_mapped_plus_unmapped(db):
    """파일 합계 = 대시보드 매출 + 미매핑. 이 항등식이 깨지면 매출이 새고 있다."""
    t, q = rp.totals(db, PERIOD), rp.quality(db, PERIOD)
    assert db.result.gross_sum == t.revenue + q.unmapped_amount


# ── T-I4 · 채널별 컬럼명이 달라도 읽힌다 ─────────────────────────────────

def test_i4_all_five_sheets_ingested(db):
    assert db.result.files == 5
    assert db.result.skipped_sheets == []
    assert db.result.lines == 3714


def test_i4b_actual_fee_wins_where_channel_provides_it(db):
    """수수료 컬럼을 주는 채널은 실적값을, 안 주는 채널은 요율을 쓴다 (PRD §5-②)."""
    from sqlalchemy import select
    from app.models import SalesLine
    src = {ch: db.scalar(select(SalesLine.fee_source)
                         .where(SalesLine.channel_id == ch, SalesLine.map_status == "MAPPED"))
           for ch in ("HANBIT", "MALL21", "SWIFT_FF", "GOODMKT", "BIDNOW")}
    assert src["HANBIT"] == src["MALL21"] == "ACTUAL"
    assert src["SWIFT_FF"] == src["GOODMKT"] == src["BIDNOW"] == "RATE"


# ── T-I5 · 결과 엑셀 ─────────────────────────────────────────────────────

def test_i5_export_has_six_sheets(db):
    from app.services import export as ex
    wb = ex.build(db, PERIOD)
    assert wb.sheetnames == ["01_요약", "02_제품별", "03_채널별",
                             "04_채널x제품", "05_적자딜", "06_미매핑"]
    assert wb["05_적자딜"].max_row - 1 == 3        # 적자 딜 3건
    assert wb["06_미매핑"].max_row - 1 == 3        # 미매핑 상품명 3종


# ── T-I6 · 잘못된 파일은 조용히 넘어가지 않는다 ──────────────────────────

def test_i6_missing_required_column_raises(tmp_path):
    """필수 컬럼이 없으면 예외. 빈 값으로 채우면 매출이 증발한다."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.title = "한빛홈쇼핑"
    wb.active.append(["주문번호", "주문일자", "상품명"])      # 수량·판매금액 없음
    wb.active.append(["A1", "2026-07-01", "무언가"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)

    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as s:
        ing.load_channels(s, DATA / "02_채널정보.xlsx")
        with pytest.raises(ing.IngestError, match="필수 컬럼 누락"):
            ing.load_sales(s, path, period=PERIOD)


# ── T-I7 · 중복 업로드 방지 ──────────────────────────────────────────────

@pytest.fixture
def fresh_db():
    """마스터만 넣은 빈 원장."""
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    with Session(engine) as s:
        ing.load_products(s, DATA / "01_상품정보.xlsx")
        ing.load_channels(s, DATA / "02_채널정보.xlsx")
        yield s


SALES = DATA / f"03_매출_{PERIOD}.xlsx"


def test_i7_same_file_twice_is_rejected(fresh_db):
    """같은 파일을 두 번 올리면 매출이 두 배가 된다 — 반드시 막아야 한다."""
    ing.load_sales(fresh_db, SALES, period=PERIOD)
    before = rp.totals(fresh_db, PERIOD).revenue

    with pytest.raises(ing.IngestError, match="이미 올린 파일"):
        ing.load_sales(fresh_db, SALES, period=PERIOD)

    assert rp.totals(fresh_db, PERIOD).revenue == before


def test_i7b_replace_keeps_totals_identical(fresh_db):
    """교체 업로드는 기존을 지우고 다시 넣으므로 합계가 그대로여야 한다."""
    ing.load_sales(fresh_db, SALES, period=PERIOD)
    before = rp.totals(fresh_db, PERIOD)

    r = ing.load_sales(fresh_db, SALES, period=PERIOD, replace=True)
    after = rp.totals(fresh_db, PERIOD)

    assert r.replaced_lines == 3714
    assert r.lines == 3714
    assert (after.revenue, after.profit, after.orders) == \
           (before.revenue, before.profit, before.orders)


def test_i7c_overlapping_orders_are_skipped(fresh_db, tmp_path):
    """파일 이름·해시가 달라도 이미 적재된 주문번호는 다시 넣지 않는다."""
    from openpyxl import load_workbook
    ing.load_sales(fresh_db, SALES, period=PERIOD)
    before = rp.totals(fresh_db, PERIOD).revenue

    wb = load_workbook(SALES)                     # 내용 동일, 해시만 다른 파일
    wb["한빛홈쇼핑"]["A1"].comment = None
    copy = tmp_path / "다시받은파일.xlsx"
    wb.save(copy)

    with pytest.raises(ing.IngestError, match="이미 들어와 있습니다"):
        ing.load_sales(fresh_db, copy, period=PERIOD)
    assert rp.totals(fresh_db, PERIOD).revenue == before


def test_i7d_sha256_is_recorded(fresh_db):
    """upload_file.sha256 이 비어 있으면 중복 검사가 동작할 수 없다."""
    from app.models import UploadFile as UF
    from sqlalchemy import select as sel
    ing.load_sales(fresh_db, SALES, period=PERIOD)
    digests = set(fresh_db.scalars(sel(UF.sha256)).all())
    assert len(digests) == 1
    assert len(digests.pop()) == 64


# ── T-I8 · 복합세트 조합 매핑 ────────────────────────────────────────────

def test_i8_composed_set_carries_every_component_cost(fresh_db):
    """복합세트는 구성 SKU를 모두 담은 새 딜로 매핑해야 원가가 맞는다.

    기존 딜(즉석밥 단품)에 붙이면 즉석국·김자반 원가가 통째로 사라진다.
    화면의 '여러 제품을 조합해 새 딜 만들기'가 이 경로다.
    """
    from app.models import Deal, DealComponent, SalesLine
    from app.services import mapping as mp
    from app.services.calc import apply_to_line, deal_unit_cogs
    from sqlalchemy import select as sel

    ing.load_sales(fresh_db, SALES, period=PERIOD)
    before = rp.quality(fresh_db, PERIOD)

    target = fresh_db.scalar(
        sel(SalesLine).where(SalesLine.map_status == "UNMAPPED",
                             SalesLine.raw_product_name.like("%홈캉스%")))
    assert target is not None, "복합세트 샘플이 실습 데이터에 있어야 한다"

    fresh_db.add(Deal(id="SET-TEST-001", channel_id=target.channel_id,
                      primary_sku_id="RICE24", label="홈캉스 세트", tier="EVENT",
                      price=target.gross_amount, effective_from="2026-01-01"))
    for sku_id, qty, gift in (("RICE24", 12, 0), ("SOUP15", 8, 0), ("GIM4", 2, 1)):
        fresh_db.add(DealComponent(deal_id="SET-TEST-001", sku_id=sku_id,
                                   qty=qty, is_gift=gift))
    fresh_db.flush()

    deal = fresh_db.get(Deal, "SET-TEST-001")
    # 760×12 + 980×8 + 1,850×2 — 증정품(김자반)도 반드시 들어간다
    assert deal_unit_cogs(fresh_db, deal, "2026-07-15") == 760 * 12 + 980 * 8 + 1850 * 2

    mp.register(fresh_db, channel_id=target.channel_id, product=target.raw_product_name,
                option=target.raw_option_name, deal_id=deal.id, match_type="COMPOSED")
    lines = fresh_db.scalars(sel(SalesLine).where(
        SalesLine.channel_id == target.channel_id,
        SalesLine.raw_product_name == target.raw_product_name,
        SalesLine.map_status == "UNMAPPED")).all()
    mp.map_lines(fresh_db, lines)
    for ln in lines:
        apply_to_line(fresh_db, ln, deal)
    fresh_db.flush()

    after = rp.quality(fresh_db, PERIOD)
    assert after.unmapped_count < before.unmapped_count
    assert lines[0].cogs == 20660
