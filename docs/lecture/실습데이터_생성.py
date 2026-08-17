# -*- coding: utf-8 -*-
"""바이브코딩 수업용 실습 파일 생성기.

    python docs/lecture/실습데이터_생성.py

DB(`data/cecnr.db`)에 들어 있는 목업 데이터를 읽어 아래를 만든다.

    실습데이터/01_상품정보.xlsx     ← 학생에게 주는 인풋 ①
    실습데이터/02_채널정보.xlsx     ← 학생에게 주는 인풋 ②
    실습데이터/03_매출_2026-07.xlsx ← 학생에게 주는 인풋 ③ (채널별 시트, 컬럼명 제각각)
    결과예시/결과_2026-07.xlsx      ← 정답지. 학생 결과물과 대조한다
    결과예시/대시보드.html          ← 정답지. 브라우저로 바로 열린다

인풋과 정답지를 같은 DB에서 뽑기 때문에 숫자가 어긋날 수 없다.
DB가 없으면 먼저:  python -m app.cli init && python -m app.demo
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook                                    # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill         # noqa: E402
from openpyxl.utils import get_column_letter                     # noqa: E402
from sqlalchemy import func, select                              # noqa: E402

from app.db import SessionLocal                                  # noqa: E402
from app.models import Channel, SalesLine                        # noqa: E402
from app.seed.data import CHANNELS, SKUS                         # noqa: E402
from app.services import export as ex                            # noqa: E402
from app.services import report as rp                            # noqa: E402

OUT_IN = Path(__file__).parent / "실습데이터"
OUT_RES = Path(__file__).parent / "결과예시"
PERIOD = "2026-07"

HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
MONEY = "#,##0"
PCT = "0.00%"

# 채널마다 컬럼명이 다르다 — 이게 4차시(파일 표준화)의 교보재다.
# (논리필드 → 그 채널에서의 실제 컬럼명). 값이 None이면 그 채널은 그 컬럼을 안 준다.
CHANNEL_COLUMNS = {
    "HANBIT":   {"주문번호": "주문번호", "주문일": "주문일자", "상품명": "상품명",
                 "옵션명": "옵션", "수량": "판매수량", "판매금액": "판매금액", "수수료": "수수료"},
    "SWIFT_FF": {"주문번호": "주문ID", "주문일": "결제일", "상품명": "노출상품명",
                 "옵션명": "등록옵션명", "수량": "수량", "판매금액": "결제금액", "수수료": None},
    "MALL21":   {"주문번호": "주문번호", "주문일": "구매확정일", "상품명": "상품명",
                 "옵션명": "단품명", "수량": "주문수량", "판매금액": "상품금액",
                 "수수료": "서비스이용료"},
    "GOODMKT":  {"주문번호": "주문번호", "주문일": "주문일", "상품명": "주문상품",
                 "옵션명": "선택옵션", "수량": "개수", "판매금액": "정산금액", "수수료": None},
    "BIDNOW":   {"주문번호": "주문번호", "주문일": "주문일", "상품명": "주문상품",
                 "옵션명": "선택옵션", "수량": "개수", "판매금액": "정산금액", "수수료": None},
}

FIELD_ORDER = ["주문번호", "주문일", "상품명", "옵션명", "수량", "판매금액", "수수료"]


# ── 시트 헬퍼 ───────────────────────────────────────────────────────────

def sheet(wb, title, headers, widths=None):
    ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
    ws.title = title
    ws.append(headers)
    for i, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = (widths or {}).get(i, 14)
    ws.freeze_panes = "A2"
    return ws


def fmt(ws, col_letters, number_format):
    for col in col_letters:
        for cell in ws[col][1:]:
            cell.number_format = number_format


# ── 인풋 ① 상품정보 ─────────────────────────────────────────────────────

def make_products():
    wb = Workbook()
    ws = sheet(wb, "상품정보",
               ["제품코드", "제품명", "규격", "매입가", "구성수량", "정상가", "일반행사가", "특가"],
               {1: 12, 2: 32, 3: 22})
    for sku_id, name, spec, cost, pack, p1, p2, p3 in SKUS:
        ws.append([sku_id, name, spec, cost, pack, p1, p2, p3])
    fmt(ws, "DFGH", MONEY)

    note = wb.create_sheet("읽어주세요")
    for r in [
        ["이 파일은 회사가 이미 가지고 있는 '가격표'입니다."], [],
        ["매입가", "제품 1개당 원가입니다. 세트 원가가 아닙니다."],
        ["구성수량", "한 번 팔 때 나가는 개수. 원가 = 매입가 × 구성수량"],
        ["", "여기를 1로 잘못 적으면 원가가 절반이 되어 이익이 두 배로 부풀려집니다."],
        ["정상가/일반행사가/특가", "같은 제품을 파는 세 가지 가격입니다. 티어라고 부릅니다."],
        [], ["⚠ 전부 가상(목업) 데이터입니다. 실존 기업·제품과 무관합니다."],
    ]:
        note.append(r)
    note.column_dimensions["A"].width = 22
    note.column_dimensions["B"].width = 80

    OUT_IN.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_IN / "01_상품정보.xlsx")
    return len(SKUS)


# ── 인풋 ② 채널정보 ─────────────────────────────────────────────────────

def make_channels():
    wb = Workbook()
    ws = sheet(wb, "채널정보",
               ["채널코드", "채널명", "그룹", "운영상태", "수수료율", "요율근거",
                "배송주체", "물류비", "물류비추정", "정산기준일", "비고"],
               {1: 12, 2: 18, 11: 46})
    for (cid, name, group, status, rate, src, ship, _m, logi, est, settle, _v, _o, note) in CHANNELS:
        ws.append([cid, name, group, status, rate, src or "", ship,
                   logi if logi is not None else "채널부담", est, settle, note])
    fmt(ws, "E", PCT)
    fmt(ws, "H", MONEY)

    note = wb.create_sheet("읽어주세요")
    for r in [
        ["채널마다 수수료가 다릅니다. 이 파일이 이 시스템의 핵심입니다."], [],
        ["수수료율", "정산서에서 역산한 값입니다.  수수료 합계 ÷ 매출 합계"],
        ["요율근거", "MEASURED=실측 / CONTRACT=계약서 / ESTIMATE=아직 못 구해 추정"],
        ["배송주체", "SELF=자사발송(물류비 우리 부담) / CHANNEL_FULFILL=채널풀필먼트 / CONSIGNMENT=위탁매입(채널 부담)"],
        ["물류비", "주문 1건당 원. '채널부담'이면 0으로 계산합니다."],
        [], ["가장 싼 채널 6.8% vs 가장 비싼 채널 24.0% — 3.5배 차이입니다."],
        ["'수수료는 대충 13%' 라는 가정이 왜 위험한지가 여기 다 나옵니다."],
        [], ["⚠ 전부 가상(목업) 데이터입니다. 실존 판매채널과 무관합니다."],
    ]:
        note.append(r)
    note.column_dimensions["A"].width = 14
    note.column_dimensions["B"].width = 100

    wb.save(OUT_IN / "02_채널정보.xlsx")
    return len(CHANNELS)


# ── 인풋 ③ 매출 원장 ────────────────────────────────────────────────────

def make_sales(db):
    wb = Workbook()
    names = {c.id: c.name for c in db.scalars(select(Channel)).all()}
    first = True
    total = 0

    for ch, cols in CHANNEL_COLUMNS.items():
        fields = [f for f in FIELD_ORDER if cols.get(f)]
        headers = [cols[f] for f in fields]
        ws = wb.active if first else wb.create_sheet()
        ws.title = names.get(ch, ch)[:31]
        first = False
        ws.append(headers)
        for i in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=i)
            c.fill, c.font = HEAD_FILL, HEAD_FONT
            ws.column_dimensions[get_column_letter(i)].width = 16 if i > 2 else 40
        ws.column_dimensions["A"].width = 14
        ws.freeze_panes = "A2"

        lines = db.scalars(
            select(SalesLine).where(SalesLine.channel_id == ch,
                                    SalesLine.period_month == PERIOD)
            .order_by(SalesLine.order_date, SalesLine.order_no)).all()
        for ln in lines:
            vals = {"주문번호": ln.order_no, "주문일": ln.order_date,
                    "상품명": ln.raw_product_name, "옵션명": ln.raw_option_name,
                    "수량": ln.qty, "판매금액": ln.gross_amount,
                    "수수료": ln.channel_fee}
            ws.append([vals[f] for f in fields])
        total += len(lines)

        money_cols = [get_column_letter(i + 1) for i, f in enumerate(fields)
                      if f in ("판매금액", "수수료")]
        fmt(ws, money_cols, MONEY)

    note = wb.create_sheet("읽어주세요")
    for r in [
        ["채널마다 시트가 따로 있고, 컬럼 이름이 전부 다릅니다."],
        ["이건 실수가 아니라 현실입니다. 채널 어드민이 각자 다르게 뽑아 줍니다."], [],
        ["같은 뜻인데 이름만 다른 것들:"],
        ["주문일", "주문일자 / 결제일 / 구매확정일 / 주문일"],
        ["상품명", "상품명 / 노출상품명 / 주문상품"],
        ["옵션명", "옵션 / 등록옵션명 / 단품명 / 선택옵션"],
        ["수량", "판매수량 / 수량 / 주문수량 / 개수"],
        ["판매금액", "판매금액 / 결제금액 / 상품금액 / 정산금액"],
        [], ["수수료 컬럼은 2개 채널만 줍니다 (한빛홈쇼핑, 몰이십일)."],
        ["나머지는 채널정보의 수수료율로 계산해야 합니다."],
        [], ["⚠ 전부 가상(목업) 데이터입니다."],
    ]:
        note.append(r)
    note.column_dimensions["A"].width = 14
    note.column_dimensions["B"].width = 70

    wb.save(OUT_IN / f"03_매출_{PERIOD}.xlsx")
    return total


# ── 정답지 ① 결과 엑셀 ──────────────────────────────────────────────────

def make_result(db):
    """앱의 내보내기 기능(app/services/export.py)을 그대로 쓴다.

    정답지와 앱의 출력이 다르면 안 되므로 로직을 복제하지 않는다.
    """
    OUT_RES.mkdir(parents=True, exist_ok=True)
    ex.build(db, PERIOD).save(OUT_RES / f"결과_{PERIOD}.xlsx")
    return rp.totals(db, PERIOD)


# ── 정답지 ② 단일 파일 대시보드 ─────────────────────────────────────────

def make_dashboard(db):
    t = rp.totals(db, PERIOD)
    q = rp.quality(db, PERIOD)
    data = {
        "period": PERIOD,
        "totals": {"revenue": t.revenue, "fee": t.fee, "cogs": t.cogs,
                   "logistics": t.logistics, "profit": t.profit,
                   "margin": t.margin, "orders": t.orders, "qty": t.qty},
        "quality": {"unmapped_count": q.unmapped_count,
                    "unmapped_amount": q.unmapped_amount,
                    "mapping_rate": q.mapping_rate},
        "channels": [{"name": r.name, "orders": r.orders, "revenue": r.revenue,
                      "fee": r.fee, "cogs": r.cogs, "logistics": r.logistics,
                      "profit": r.profit, "margin": r.margin,
                      "rate": r.extra.get("fee_rate")} for r in rp.by_channel(db, PERIOD)],
        "products": [{"name": r.name, "qty": r.qty, "revenue": r.revenue,
                      "profit": r.profit, "margin": r.margin}
                     for r in rp.by_product(db, PERIOD)],
        "losses": [{"channel": r.sub, "name": r.name, "tier": r.extra["tier"],
                    "price": r.extra["price"], "orders": r.orders,
                    "revenue": r.revenue, "profit": r.profit, "margin": r.margin}
                   for r in rp.loss_deals(db, PERIOD)],
        "weeks": [{"week": r.name, "revenue": r.revenue, "profit": r.profit,
                   "margin": r.margin} for r in rp.weekly_trend(db, PERIOD)],
    }
    tpl = (Path(__file__).parent / "_대시보드_템플릿.html").read_text(encoding="utf-8")
    html = tpl.replace("/*__DATA__*/null",
                       json.dumps(data, ensure_ascii=False, separators=(",", ":")))
    OUT_RES.mkdir(parents=True, exist_ok=True)
    (OUT_RES / "대시보드.html").write_text(html, encoding="utf-8")


if __name__ == "__main__":
    with SessionLocal() as db:
        if not db.scalar(select(func.count()).select_from(SalesLine)):
            sys.exit("판매 데이터가 없습니다. 먼저: python -m app.cli init && python -m app.demo")
        n_sku, n_ch = make_products(), make_channels()
        n_line = make_sales(db)
        t = make_result(db)
        make_dashboard(db)

    print("실습 파일 생성 완료\n")
    print(f"  인풋   실습데이터/01_상품정보.xlsx        제품 {n_sku}개")
    print(f"         실습데이터/02_채널정보.xlsx        채널 {n_ch}개")
    print(f"         실습데이터/03_매출_{PERIOD}.xlsx    주문 {n_line:,}건 (채널별 시트 5장)")
    print(f"  정답지 결과예시/결과_{PERIOD}.xlsx        시트 6장")
    print(f"         결과예시/대시보드.html             브라우저로 열면 바로 보임")
    print(f"\n  검산: 총매출 {t.revenue:,}원 · 기여이익 {t.profit:,}원 · 마진율 {t.margin:.2%}")
